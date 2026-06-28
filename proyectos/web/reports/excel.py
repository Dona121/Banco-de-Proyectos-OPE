"""Generación de archivos Excel (.xlsx) de los reportes."""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

VERDE = "FF109D39"
VERDE_CLARO = "FFE7F6EC"
GRIS = "FF5A595D"

_titulo = Font(name="Calibri", size=15, bold=True, color=VERDE)
_meta = Font(name="Calibri", size=9, color=GRIS)
_th = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
_fill_th = PatternFill("solid", fgColor=VERDE)
_fill_alt = PatternFill("solid", fgColor=VERDE_CLARO)
_center = Alignment(horizontal="center", vertical="center")
_left = Alignment(horizontal="left", vertical="center")
_thin = Side(style="thin", color="FFD9E2EC")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def proyectos_formulados_xlsx(filas, *, generado_por, generado_en):
    """Devuelve los bytes de un .xlsx con los proyectos formulados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Proyectos formulados"

    columnas = [
        ("Proyecto", 42, "left"),
        ("Responsable", 26, "left"),
        ("Total actividades", 16, "center"),
        ("Aprobadas", 12, "center"),
        ("Fecha creación", 18, "center"),
        ("Última actualización", 20, "center"),
        ("Estado", 14, "center"),
    ]

    # Encabezado institucional.
    ws.merge_cells("A1:G1")
    ws["A1"] = "Gobernación de Sucre · Reporte de Proyectos Formulados"
    ws["A1"].font = _titulo
    ws.merge_cells("A2:G2")
    ws["A2"] = (
        f"Generado por {generado_por} · {generado_en:%d/%m/%Y %H:%M} · "
        f"{len(filas)} proyecto(s)"
    )
    ws["A2"].font = _meta

    fila_th = 4
    for i, (titulo, ancho, _) in enumerate(columnas, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = ancho
        c = ws.cell(row=fila_th, column=i, value=titulo)
        c.font = _th
        c.fill = _fill_th
        c.alignment = _center
        c.border = _border

    for r, fila in enumerate(filas, start=fila_th + 1):
        valores = [
            fila["proyecto"].nombre,
            (fila["responsable"].get_full_name() or fila["responsable"].username),
            fila["total"],
            fila["aprobadas"],
            fila["fecha_creacion"].strftime("%d/%m/%Y"),
            fila["fecha_actualizacion"].strftime("%d/%m/%Y"),
            fila["estado"],
        ]
        for i, (valor, (_, _, align)) in enumerate(zip(valores, columnas), start=1):
            c = ws.cell(row=r, column=i, value=valor)
            c.alignment = _center if align == "center" else _left
            c.border = _border
            if (r - fila_th) % 2 == 0:
                c.fill = _fill_alt

    ws.freeze_panes = ws.cell(row=fila_th + 1, column=1)
    ws.auto_filter.ref = f"A{fila_th}:G{fila_th + len(filas)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
